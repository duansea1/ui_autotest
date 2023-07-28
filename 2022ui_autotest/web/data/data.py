# -- coding: utf-8 --
# @time :
# @author : xxxx
# @file : .py
# @desp : 加载框架数据
import csv
from pathlib import Path
from openpyxl import load_workbook


data_dir = Path(__file__).parent.parent / "data"  # 数据存放目录


def read_csv(filename):

    path = data_dir / filename
    with open(path, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for i in reader:
            print("i:", i.values())
            yield list(i.values())  # 重点返回参数，未存储


def read_excel(filename):
    path = data_dir / filename
    wb = load_workbook(path)
    ws = wb.active
    yield from ws.iter_rows(min_row=2,
                            values_only=True)


if __name__ == "__main__":
    for d in read_excel(filename="ddt_test_new_adress.xlsx"):
        print("d:", d)