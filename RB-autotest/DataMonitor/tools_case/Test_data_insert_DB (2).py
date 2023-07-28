'''
# -*- coding: utf-8 -*-
#开发人员:   chenpeng
#开发日期:   2019-07-01
#文件项目:   测试数据维护
#文件名称:   测试数据维护到数据库
 '''
from public.mysql_opr import query_mysql
import os
import openpyxl
DB_info = {"host" : "yyw-0345","user" : "qauser","password" : "qa_123456","port" : 3306,"data" : "qateam"}
Root_path = os.path.abspath (os.path.join (os.getcwd (), ".."))  # 获取AutoTest目录
# print(Root_path)
test_data_excel_path = os.path.join (Root_path, '自动化测试账号.xlsx')  # 拼接测试数据文件路径
print (test_data_excel_path)
wb = openpyxl.load_workbook (test_data_excel_path)  # 打开已有的excel
sheet_names = wb.sheetnames  # 得到工作簿的所有工作表名
user_data = wb[sheet_names[0]]  # 打开第一个 sheet 工作表
user_qty = int (user_data.title)  # 当前账号信息条数
print (user_qty)
sheet1_max_row = user_data.max_row  # 获取最大行数 结果：10
print (sheet1_max_row)
if user_qty != sheet1_max_row :
    user_data.title = str (sheet1_max_row)  # 将sheet重命名为最大行数
    wb.save (test_data_excel_path)
def query_excel_data_every_row(row_num):
    '''取出新增的数据'''
    row_data_list = []
    for i in user_data[row_num]:
        row_data_list.append (str(i.value))
    row_data_list=row_data_list[0:-2] #取出前9位数据
    row_data_tuple=tuple(row_data_list)
    print('元组：',row_data_tuple)
    return row_data_tuple

def insert_data_to_DB():
    '''判断excel数据是否更新，如有更新则插入数据,暂不支持更新已有数据'''
    if user_qty < sheet1_max_row:
        for a in range (user_qty+1, sheet1_max_row + 1) :
            row_data = query_excel_data_every_row (a)
            #print('111',row_data)
            Insert_sql="INSERT INTO auto_account_records (`platform`, `userid`, `username`, `password`,`entername`, `phone`, `description`, `environment`, `protect_type`)" \
                       " VALUES %s" %(row_data,) #必须➕逗号 不然sql错误。。
            print(Insert_sql)
            query_mysql (DB_info['host'], DB_info['user'], DB_info['password'], DB_info['port'], DB_info['data'],sql=Insert_sql)
    else:
        print('当前无账号信息更新')


if __name__ == '__main__':
    pass
    insert_data_to_DB()






