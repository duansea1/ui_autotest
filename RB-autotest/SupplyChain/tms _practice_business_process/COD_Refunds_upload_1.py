#coding=utf-8
from selenium import webdriver
from time import sleep
import os
import openpyxl
import random
import mysql.connector
'''未导入订单导入返款数据'''
browser = webdriver.Chrome()
browser.maximize_window() #最大化浏览器
url='http://tms2.111.com.cn/home'
print ("进入%s"%(url))#进入TMS
browser.get(url)
# login()
print("输入用户名")
browser.find_element_by_name("username").send_keys("chenpeng")
print("输入密码")
browser.find_element_by_name("password").send_keys("Cp,456789")
print("点击登录")
browser.find_element_by_css_selector("#login-from > span").click()
sleep(3)
title=browser.title
print("页面标题为："+title)
url=browser.current_url
print("当前网址为："+url)
sleep(3)
farther=browser.find_elements_by_css_selector("ul.layui-layout-left>li>a")
son=browser.find_elements_by_css_selector("li.layui-nav-item>dl>dd>a")
sleep(2)
first='费用结算'
second='COD订单管理'
for i in farther:
    if i.text.strip() ==first:
        print("点击一级菜单："+first)
        i.click()
        sleep(1)
        for a in son:
            if a.text.strip()==second:
                print("点击二级菜单："+second)
                a.click()
                sleep(1)
frame=browser.find_element_by_css_selector("body > div > div.layui-body > div > div > div > div > iframe")
print("切换frame")
browser.switch_to.frame(frame)
sleep(1)
browser.find_element_by_xpath('//*[@id="ra"]/div[2]/div/div/div/input').click()
sleep(1)
# print('选择全部')
# browser.find_element_by_xpath('//*[@id="ra"]/div[2]/div/div/dl/dd[1]').click()
print('选择未导入订单')
browser.find_element_by_xpath('//*[@id="ra"]/div[2]/div/div/dl/dd[2]').click()
sleep(1)
print("输入出库时间")
browser.find_element_by_id('leaveDcTimeStart').clear()
sleep(1)
browser.find_element_by_id('leaveDcTimeStart').send_keys('2017-04-01 00:00:00')
browser.find_element_by_id("btn-submit").click()#点击查询
sleep(2)
do=browser.find_element_by_xpath('/html/body/div/div/div/div[8]/div[1]/div[2]/table/tbody/tr/td[4]/div').text
carrier=browser.find_element_by_xpath('/html/body/div/div/div/div[8]/div[1]/div[2]/table/tbody/tr/td[7]/div').text
#-----------------------------------------------------------------
'''连接数据库查询承运商code'''
mydb = mysql.connector.connect(
                host="10.6.84.20",
                user="tms",
                passwd="d41d8cd98f00b204",
                database="tms",
)
mycursor = mydb.cursor()
query='SELECT code from tms_carrier where name=%s'
mycursor.execute(query,(carrier,))
myresult = mycursor.fetchall()
b=myresult[0]
carrie_code = ','.join(b)
#----------------------------------------------------------
refunds1=browser.find_element_by_xpath('/html/body/div[1]/div/div/div[8]/div[1]/div[2]/table/tbody/tr/td[15]/div').text
status1=browser.find_element_by_xpath('/html/body/div[1]/div/div/div[8]/div[1]/div[2]/table/tbody/tr/td[17]/div').text
print('当前DO：'+do)
print('当前承运商CODE：'+carrie_code)
print('当前返款金额：'+refunds1)
print('当前返款状态：'+status1)
sleep(2)
browser.find_element_by_xpath('//*[@id="uploadExcel"]/button').click()
sleep(5)
#----------------------------------------------------------
a_path = (os.path.abspath (os.path.join (os.getcwd (), "..")))  # 获取SuppluChain目录
print(a_path)
excel_path = os.path.join (a_path, r"file\tms\cod_refunds.xlsx")  # 拼接出导入文件路径
print(excel_path)
exe_path = os.path.join (a_path, r"file\tms\cod_refunds_1.exe %s")  # 拼接出导入文件路径
print(exe_path)
#os.system (exe_path % excel_path)  # 对excel文件路径参数化执行exe上传文件
# os.system(r"D:\9\test-autotest\SupplyChain\file\tms\delivery_time_template_1.exe %s" % excel_path)  #绝对路径也可以实现

#----------------------------------------------------------


print('--------------修改Excel返款金额开始-----------')
# wb=openpyxl.load_workbook(r'C:\Users\chenpeng\Desktop\test_file\cod_refunds.xlsx')
wb=openpyxl.load_workbook(excel_path)

print(wb.sheetnames)
sheet=wb['返款对账']
i=random.randint(1,100)
sheet['C2'] =i  # 修改返款金额
sheet['B2'] =do  # 修改承运商
sheet['A2'] =carrie_code  # 修改do
#wb.save(r'C:\Users\chenpeng\Desktop\test_file\cod_refunds.xlsx')
wb.save(excel_path)

print('--------------修改Excel返款金额完成-----------')
sleep(1)
print('--------------上传返款数据开始----------------')
#os.system(r'C:\Users\chenpeng\Desktop\test_file\upload.exe')
os.system (exe_path % excel_path)  # 对excel文件路径参数化执行exe上传文件

print('--------------上传返款数据完成----------------')
sleep(3)
browser.find_element_by_name('code').send_keys(do)
sleep(1)
browser.find_element_by_xpath('//*[@id="ra"]/div[2]/div/div/div/input').click()
sleep(1)
print('选择全部')
browser.find_element_by_xpath('//*[@id="ra"]/div[2]/div/div/dl/dd[1]').click()
browser.find_element_by_id("btn-submit").click()#点击查询
sleep(3)
print('勾选订单')
browser.find_element_by_xpath('/html/body/div[1]/div/div/div[8]/div[1]/div[3]/div[2]/table/tbody/tr[1]/td/div/div/i').click()
sleep(1)
print ('点击返款')
browser.find_element_by_id ('confirm').click ()
sleep(1)
refunds2=browser.find_element_by_xpath('/html/body/div[1]/div/div/div[8]/div[1]/div[2]/table/tbody/tr/td[15]/div').text
status2=browser.find_element_by_xpath('/html/body/div[1]/div/div/div[8]/div[1]/div[2]/table/tbody/tr/td[17]/div').text
collect_amout=browser.find_element_by_xpath('/html/body/div/div/div/div[8]/div[1]/div[2]/table/tbody/tr/td[14]/div').text
print('应收金额：'+collect_amout)
print('导入返款金额：'+refunds2)
print('导入返款状态：'+status2)
if refunds1==refunds2:
    if status2 =="已返款":
        print('返款成功，测试通过')
elif refunds1 != refunds2:
    if status2 == '异常':
        print('返款异常，测试通过')
else:
    if status2 == '未导入' or status2 == '处理中':
        print('返款失败，测试不通过')